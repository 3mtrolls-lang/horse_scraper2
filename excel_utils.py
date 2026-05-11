"""
excel_utils.py - Export race records to formatted Excel workbooks.

Creates two workbooks:
  1. latest_run_TIMESTAMP.xlsx  — single most recent / nearest upcoming race
  2. history_TIMESTAMP.xlsx     — full race history

Both use the same column schema defined in config.EXCEL_COLUMNS.
A third sheet "Reference" is added to both files with:
  - Centre/Track code mappings
  - Equipment code mappings
"""

import os
import logging
from datetime import datetime
from typing import Any

import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from config import (
    EXCEL_COLUMNS,
    OUTPUT_DIR,
    TRACK_CODE_MAP,
    EQUIPMENT_CODE_MAP,
    DATE_FORMAT,
)
from models import RaceRecord

logger = logging.getLogger("horse_scraper.excel_utils")

# ─── Header style ─────────────────────────────────────────────
_HEADER_FILL  = PatternFill("solid", fgColor="1F3864")   # dark navy
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_ALT_FILL     = PatternFill("solid", fgColor="EBF0FA")   # light blue-grey
_BORDER_THIN  = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

_DATE_NUMFMT  = "DD-MM-YYYY"
_NUM_COLS     = {
    "WeightAllotted", "WeightCarried", "Wt",
    "Pre_Rating_Official", "Post_Rating_Official", "Turf Rating",
    "Starters", "RaceNo", "H No", "D No",
}


def _records_to_df(records: list[RaceRecord]) -> pd.DataFrame:
    """Convert list of RaceRecord → DataFrame with exact column order."""
    rows = [r.to_dict() for r in records]
    df = pd.DataFrame(rows, columns=EXCEL_COLUMNS)
    # Replace any None / NaN with empty string
    df = df.fillna("").infer_objects(copy=False)
    return df


def _style_worksheet(ws, df: pd.DataFrame, sheet_title: str = "") -> None:
    """Apply header styles, alternating row fills, column widths, freeze panes."""
    # ── Header row ──────────────────────────────────────────────
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _BORDER_THIN

    # ── Data rows ───────────────────────────────────────────────
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        fill = _ALT_FILL if row_idx % 2 == 0 else PatternFill()
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            col_name = df.columns[col_idx - 1]

            # Numeric columns
            if col_name in _NUM_COLS and value not in ("", None):
                try:
                    cell.value = float(str(value))
                    cell.number_format = "0.00" if "." in str(value) else "0"
                except ValueError:
                    cell.value = value
            else:
                cell.value = str(value) if value not in ("", None) else ""

            cell.fill = fill
            cell.border = _BORDER_THIN
            cell.alignment = Alignment(vertical="center")

    # ── Auto column widths (max 40, min 8) ──────────────────────
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(col_name))
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 40)

    # ── Freeze header ────────────────────────────────────────────
    ws.freeze_panes = ws["A2"]

    # ── Auto-filter ──────────────────────────────────────────────
    ws.auto_filter.ref = ws.dimensions

    # ── Row height ───────────────────────────────────────────────
    ws.row_dimensions[1].height = 32
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 18


def _add_reference_sheet(wb: openpyxl.Workbook) -> None:
    """Add a 'Reference' sheet with code mappings."""
    ws = wb.create_sheet("Reference")

    # Track codes
    ws["A1"] = "TRACK / CENTRE CODES"
    ws["A1"].font = Font(bold=True, size=11, color="1F3864")

    ws["A2"] = "Full Name"
    ws["B2"] = "Code"
    for cell in [ws["A2"], ws["B2"]]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN

    for i, (full, code) in enumerate(sorted(TRACK_CODE_MAP.items()), start=3):
        ws.cell(row=i, column=1, value=full.title())
        ws.cell(row=i, column=2, value=code)
        if i % 2 == 0:
            for c in [ws.cell(row=i, column=1), ws.cell(row=i, column=2)]:
                c.fill = _ALT_FILL

    # Equipment codes — offset to column D
    offset_row = 2
    ws.cell(row=1, column=4, value="EQUIPMENT CODES").font = Font(bold=True, size=11, color="1F3864")
    ws.cell(row=offset_row, column=4, value="Equipment").font = _HEADER_FONT
    ws.cell(row=offset_row, column=4).fill = _HEADER_FILL
    ws.cell(row=offset_row, column=4).alignment = _HEADER_ALIGN
    ws.cell(row=offset_row, column=5, value="Code").font = _HEADER_FONT
    ws.cell(row=offset_row, column=5).fill = _HEADER_FILL
    ws.cell(row=offset_row, column=5).alignment = _HEADER_ALIGN

    for i, (equip, code) in enumerate(sorted(EQUIPMENT_CODE_MAP.items()), start=offset_row + 1):
        ws.cell(row=i, column=4, value=equip.title())
        ws.cell(row=i, column=5, value=code)
        if i % 2 == 0:
            for c in [ws.cell(row=i, column=4), ws.cell(row=i, column=5)]:
                c.fill = _ALT_FILL

    # Auto-width columns A, B, D, E
    for col_letter, min_w in [("A", 30), ("B", 8), ("D", 30), ("E", 10)]:
        ws.column_dimensions[col_letter].width = min_w

    ws.freeze_panes = "A3"


def _get_timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def export_latest_run(records: list[RaceRecord]) -> str:
    """
    Export the latest run (most recent completed OR nearest upcoming) records.
    One row per horse.

    Returns the filepath of the saved Excel file.
    """
    ts = _get_timestamp()
    path = os.path.join(OUTPUT_DIR, f"latest_run_{ts}.xlsx")

    df = _records_to_df(records)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Latest Run"
    _style_worksheet(ws, df, "Latest Run")
    _add_reference_sheet(wb)

    # Add a second sheet "Upcoming" filtered to UPCOMING races
    upcoming = df[df["RaceStatus"] == "UPCOMING"].copy()
    if not upcoming.empty:
        ws_up = wb.create_sheet("Upcoming Entries")
        _style_worksheet(ws_up, upcoming, "Upcoming Entries")

    wb.save(path)
    logger.info("Latest run Excel saved → %s", path)
    return path


def export_history(records: list[RaceRecord]) -> str:
    """
    Export the full race history records (all runs for all horses).

    Returns the filepath of the saved Excel file.
    """
    ts = _get_timestamp()
    path = os.path.join(OUTPUT_DIR, f"history_{ts}.xlsx")

    df = _records_to_df(records)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Race History"
    _style_worksheet(ws, df, "Race History")
    _add_reference_sheet(wb)

    # Per-horse summary sheet
    if "Horse Name" in df.columns and not df.empty:
        summary_rows = []
        for horse, grp in df.groupby("Horse Name"):
            summary_rows.append({
                "Horse Name": horse,
                "Total Runs": len(grp),
                "Wins": (grp["Placing2"].astype(str) == "1").sum(),
                "Earliest Race": grp["RaceDate"].min(),
                "Latest Race": grp["RaceDate"].max(),
                "Country": grp["CountryCode"].iloc[0] if "CountryCode" in grp.columns else "",
            })
        summary_df = pd.DataFrame(summary_rows)
        ws_sum = wb.create_sheet("Summary")
        _style_worksheet(ws_sum, summary_df, "Summary")

    wb.save(path)
    logger.info("History Excel saved → %s", path)
    return path
